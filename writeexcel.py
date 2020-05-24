import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active
print(sheet.title)
sheet.title = '2BDC'
print(sheet.title)
print(wb.sheetnames)
wb.create_sheet(index=1,title='6BDC')
wb.create_sheet(index=2,title='Fab2 Corp')
wb.create_sheet(index=3,title='Fab2 Dev')
wb.create_sheet(index=4,title='Fab7 Corp')
wb.create_sheet(index=5,title='EBIZ')
wb.create_sheet(index=5,title='ERP')
wb.create_sheet(index=5,title='ERP-DR')
print(wb.sheetnames)

wb.save('Ip address and Hostnames.xlsx')

wb = openpyxl.load_workbook('Ip address and Hostnames.xlsx')
sheet = wb['2BDC']

for i in range(1,11):
    sheet['A{}'.format(i)]= i
    sheet['B{}'.format(i)]= i

wb.save('Ip address and Hostnames.xlsx')
