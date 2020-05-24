import openpyxl
import ipaddress
import subprocess
import socket
from subprocess import Popen, PIPE
wb = openpyxl.load_workbook(r"C:\Users\RamuGajula\Desktop\GF-V2\IP Inventory v1.0.xlsx")
ipinventory = wb['Sheet1']
for i in range(ipinventory.max_row):
    # print(ipinventory.cell(row = i+1,column = 7).value)
    netaddr = ipinventory.cell(row = i+1,column = 8).value
    try:

        ip_net = ipaddress.ip_network(netaddr)
        print("Initiating ping on Subnet:",netaddr)

    except ValueError:
       continue
    #print(ip_net.num_addresses)

    for j in ip_net.hosts():
        ip_add = str(j)
        #print(ip_add)
        res = subprocess.call(['ping','-n','1',ip_add],stdout=PIPE)
        if res == 0:
            #rev_name = socket.gethostbyadd(j)
            try:
             rev_name = socket.gethostbyaddr(ip_add)
            except socket.herror:
             print(ip_add,"is alive but FQDN is not found")
             continue
            print(ip_add,"is alive and FQDN is "+rev_name[0])
        else:
            print(ip_add,"is not alive")

    # print("========="+str(i)+"==========")
