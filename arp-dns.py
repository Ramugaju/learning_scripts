import openpyxl
import ipaddress
import subprocess
import socket
from subprocess import Popen, PIPE
wb = openpyxl.load_workbook(r"C:\Users\RamuGajula\Desktop\GF-V2\IP Inventory v1.0.xlsx")
ipinventory = wb['ARP Table']
#for i in range(ipinventory.max_row):
wb = openpyxl.Workbook()
sheet = wb.active
print(sheet.title)
sheet.title = '2BDC'
for i in range(0,215):
     #print(ipinventory.cell(row = i+1,column = 2).value)
     ip_add = ipinventory.cell(row = i+1,column = 2).value
     try:

        arp_entry = ipaddress.ip_address(ip_add)
        #print("Initiating ping on Subnet:",netaddr)

     except ValueError:
       continue
     print(ip_add)

    # for j in ip_net.hosts():
    #     ip_add = str(j)
    #     #print(ip_add)
    #     res = subprocess.call(['ping','-n','1',ip_add],stdout=PIPE)
    #     if res == 0:
    #         #rev_name = socket.gethostbyadd(j)
     try:
            print("Trying to resolve...",ip_add)
            rev_name = socket.gethostbyaddr(ip_add)
     except socket.herror:
             print("No hostname found for",ip_add)
             sheet['A{}'.format(i)] = ip_add
             sheet['B{}'.format(i)] = "N/A"
             continue
     sheet['A{}'.format(i)] = ip_add
     sheet['B{}'.format(i)] = rev_name[0]
     print("Hostname found "+rev_name[0]+"Entry Added Ok!")

wb.save('2BDC-ARP-DNS.xlsx')

        # else:
        #     print(ip_add,"is not alive")

    # print("========="+str(i)+"==========")
