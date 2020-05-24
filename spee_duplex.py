import re
import sys
import openpyxl
wb = openpyxl.load_workbook(r"C:\\Users\\RamuGajula\\Desktop\\Py-portmapping\\Port-Mapping.xlsx")
sheet = wb['Fab7-corp-1']
device_file_name = sys.argv[1]
f = open("C:\\Users\\RamuGajula\\Desktop\\Py-portmapping\\show int status\\{}".format(device_file_name),"r")
Interface ={}
j = 0
count_of_lines = []
for line in f:
    if "Interface" in line:
        m = re.findall(r"(.*)\s+(['connected','notconnect','disabled','monitoring','sfpAbsent','down','noOperMem']+)\s+([a-z,A-Z,0-9,/,\.]+)\s+([a-z,A-Z,0-9,/,\.,-]+)\s+([a-z,A-Z,0-9,/,\.,-]+)\s+([a-z,A-Z,0-9,/,\.]+)",line)
        # print(m[0][0])

        # for i in mo[0]:
        #     print(i)
        continue
    else:
        # print(line)
            if re.findall(r"(.*)\s+(['connected','notconnect','disabled','monitoring','sfpAbsent','down','noOperMem']+)\s+([a-z,A-Z,0-9,/,\.]+)\s+([a-z,A-Z,0-9,/,\.,-]+)\s+([a-z,A-Z,0-9,/,\.,-]+)\s+([a-z,A-Z,0-9,/,\.]+)",line):

                if "show int status" in line:
                    break
                else:
                    m = re.findall(r"(.*)\s+(['connected','notconnect','disabled','monitoring','sfpAbsent','down','noOperMem']+)\s+([a-z,A-Z,0-9,/,\.]+)\s+([a-z,A-Z,0-9,/,\.,-]+)\s+([a-z,A-Z,0-9,/,\.,-]+)\s+([a-z,A-Z,0-9,/,\.]+)",line)
                    count_of_lines.append(m[0][2])

                    # Interface[m[0][0]] = {'Ip-address':m[0][1],'OK?':m[0][2],'Method':m[0][3],'Status':m[0][4],'Protocol':m[0][5]}
                    # print(m)
                    # print(m(3))
                    # print(m(4))
                    # print(m(5))
                    if  m[0][2] == 'trunk':
                        sheet['I{}'.format(j+1)]= 'trunk'
                    elif m[0][2] == 'routed':
                        sheet['I{}'.format(j+1)]= 'NA'
                    else:
                        sheet['I{}'.format(j+1)]= 'Access'




                    sheet['J{}'.format(j+1)]= m[0][2]
                    sheet['K{}'.format(j+1)]= m[0][4]
                    sheet['L{}'.format(j+1)]= m[0][3]

                    # sheet['E{}'.format(j+2)]= m[0][4]
                    # sheet['F{}'.format(j+2)]= m[0][5]
                    # sheet['H{}'.format(j+2)]= m[0][1]
                    # print(sheet['E{}'.format(j+2)].value)

    j = j+1
# print(len(count_of_lines))
# print(len(Interface))
# print(m[0][0])
# wb = openpyxl.Workbook()

# sheet = wb.active
# print(sheet.title)
# sheet.title = device_file_name
# wb.create_sheet(title=device_file_name)
# sheet = wb['Fab7-corp-1']
# for i in range(len(Interface)):
#     sheet['C{}'.format(i)]= m[0][0]
#      #print(ipinventory.cell(row = i+1,column = 2).value)
#      # ip_add = ipinventory.cell(row = i+1,column = 2).value
#      # # try:
#      # #
#      # #    arp_entry = ipaddress.ip_address(ip_add)
#      # #    #print("Initiating ping on Subnet:",netaddr)
#      # #
#      # # except ValueError:
#      # #   continue
#      # # print(ip_add)
#
#     # for j in ip_net.hosts():
#     #     ip_add = str(j)
#     #     #print(ip_add)
#     #     res = subprocess.call(['ping','-n','1',ip_add],stdout=PIPE)
#     #     if res == 0:
#     #         #rev_name = socket.gethostbyadd(j)
#      # try:
#      #        print("Trying to resolve...",ip_add)
#      #        rev_name = socket.gethostbyaddr(ip_add)
#      # except socket.herror:
#      #         print("No hostname found for",ip_add)
#      #         sheet['A{}'.format(i)] = ip_add
#      #         sheet['B{}'.format(i)] = "N/A"
#      #         continue
#      # sheet['A{}'.format(i)] = ip_add
#      # sheet['B{}'.format(i)] = rev_name[0]
#      # print("Hostname found "+rev_name[0]+"Entry Added Ok!")
#
wb.save("C:\\Users\\RamuGajula\\Desktop\\Py-portmapping\\Port-Mapping.xlsx")
